import pandas as pd
import openpyxl
import os

def apply_npc_styling(writer, sheet_name, df):
    """Meticulous Formatting for Data Sheets."""
    workbook = writer.book
    ws = writer.sheets[sheet_name]
    header_fmt = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1, 'align': 'center'})
    right_fmt = workbook.add_format({'align': 'right'})
    center_fmt = workbook.add_format({'align': 'center'})
    int_right_fmt = workbook.add_format({'num_format': '0', 'align': 'right'})
    int_center_fmt = workbook.add_format({'num_format': '0', 'align': 'center'})

    ws.set_column('A:E', 10); ws.set_column('F:F', 16); ws.set_column('G:G', 20)
    ws.set_column('H:H', 8); ws.set_column('I:I', 11)

    for col_num, value in enumerate(df.columns):
        ws.write(0, col_num, value, header_fmt)

    ws.freeze_panes(1, 0)
    ws.autofilter(0, 0, len(df), len(df.columns) - 1)

    def safe_suffix(val, suffix):
        s_val = str(val).strip()
        return s_val if s_val.endswith(suffix.strip()) else f"{s_val} {suffix.strip()}"

    for i, row in enumerate(df.values):
        if pd.isna(row[0]) or row[0] == "": continue 
        ws.write(i + 1, 0, f"{str(row[0]).strip().rstrip(':')}: ", right_fmt)
        ws.write(i + 1, 1, row[1])
        ws.write(i + 1, 2, safe_suffix(row[2], " Years"), int_right_fmt)
        ws.write(i + 1, 3, safe_suffix(row[3], " lbs."), right_fmt)
        h = str(row[4]).strip()
        ws.write(i + 1, 4, h if h.endswith('"') else f'{h}"', center_fmt)
        ws.write(i + 1, 5, f" {row[5]}")
        ws.write(i + 1, 6, row[6], center_fmt)
        pwr_clean = str(row[7]).split()[0]
        ws.write(i + 1, 7, f"{pwr_clean} pts.", int_center_fmt)
        ws.write(i + 1, 8, safe_suffix(row[8], " Years"), int_right_fmt)

def calculate_internal_matrix(df):
    """Deep Copy Sandbox: Prevents P_Val/R_Clean from leaking to campaign.py."""
    audit_df = df.copy()
    audit_df['P_Val'] = pd.to_numeric(audit_df.iloc[:, 7].astype(str).str.split().str[0], errors='coerce').fillna(0)
    audit_df['R_Clean'] = audit_df.iloc[:, 0].astype(str).str.rstrip(': ').str.strip()
    
    races = ["Dwarf", "Elf", "Gnome", "Half-Elf", "Halfling", "Human"]
    ranks = ["Commoner", "Hero", "God"]
    internal = {rank: {race: 0 for race in races} for rank in ranks}
    
    for _, row in audit_df.iterrows():
        pwr, race = row['P_Val'], row['R_Clean']
        if race in races:
            p_cat = "God" if pwr >= 92 else ("Hero" if pwr >= 78 else "Commoner")
            internal[p_cat][race] += 1
    return internal

def export_npc_pop(lists, gen_name, matrix, folder_path=None):
    """Plugin Entry Point: Creates meticulous files and perfect summaries."""
    headers = ['Race', 'Sex', 'Age', 'Weight', 'Height', 'Alignment', 'Stats', 'Power', 'Death']
    raw_data = lists.get(gen_name, [])
    # Slice to 9 columns to guarantee no 'ValueError' in the main loop
    df_main = pd.DataFrame(raw_data).iloc[:, :9]
    df_main.columns = headers
    
    active_total = len(df_main.dropna(how='all'))
    current_matrix = calculate_internal_matrix(df_main)

    if not folder_path:
        folder_path = os.path.join("data/campaign", gen_name, "npc_population")
    os.makedirs(folder_path, exist_ok=True)

    for suffix in ["Init", "Prime"]:
        file_path = os.path.join(folder_path, f"{gen_name}_{suffix}.xlsx")
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            writer.book.add_worksheet(suffix) 
            apply_npc_styling(writer, suffix, df_main)

    summary_template_path = "data/_master_summary.xlsx"
    if os.path.exists(summary_template_path):
        wb = openpyxl.load_workbook(summary_template_path)
        ws_sum = wb.active
        ws_sum.merge_cells("A1:O1")
        ws_sum["A1"].value = gen_name
        
        row_map = {"Commoner": 3, "Hero": 4, "God": 5, "Total": 6}
        races = ["Dwarf", "Elf", "Gnome", "Half-Elf", "Halfling", "Human"]
        for p_key, target_row in row_map.items():
            #######################################################################
            # Calculate integer row sum for column 14 (N)
            row_sum = active_total if p_key == "Total" else sum(current_matrix[p_key][r] for r in races)
            
            for i, r_key in enumerate(races):
                count_col = 2 + (i * 2) # Targets B, D, F, H, J, L
                
                # Get raw integer count
                val = sum(current_matrix[p][r_key] for p in ["Commoner", "Hero", "God"]) if p_key == "Total" else current_matrix[p_key][r_key]
                
                # Write ONLY integer; Python no longer overwrites percentage formula columns
                ws_sum.cell(row=target_row, column=count_col).value = int(val)
            
            # Write integer total to column 14 (N); column 15 (O) is handled by template formulas
            ws_sum.cell(row=target_row, column=14).value = int(row_sum)
            #######################################################################
            # row_sum = active_total if p_key == "Total" else sum(current_matrix[p_key][r] for r in ["Dwarf", "Elf", "Gnome", "Half-Elf", "Halfling", "Human"])
            # for i, r_key in enumerate(["Dwarf", "Elf", "Gnome", "Half-Elf", "Halfling", "Human"]):
            #     count_col, perc_col = 2 + (i * 2), 3 + (i * 2)
            #     val = sum(current_matrix[p][r_key] for p in ["Commoner", "Hero", "God"]) if p_key == "Total" else current_matrix[p_key][r_key]
            #     ws_sum.cell(row=target_row, column=count_col).value = val
            #     ws_sum.cell(row=target_row, column=perc_col).value = (val / active_total) if active_total > 0 else 0
            #     ws_sum.cell(row=target_row, column=perc_col).number_format = '0.00%'
            # ws_sum.cell(row=target_row, column=14).value = row_sum
            # ws_sum.cell(row=target_row, column=15).value = (row_sum / active_total) if active_total > 0 else 0
            # ws_sum.cell(row=target_row, column=15).number_format = '0.00%'
        wb.save(os.path.join(folder_path, f"{gen_name}_Summary.xlsx"))

    return {k: len(v) for k, v in lists.items()}, df_main